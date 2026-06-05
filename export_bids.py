"""추천 공고 JSON → 엑셀 변환"""

import os
import json
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def json_to_excel(json_file: str) -> str:
    with open(json_file, "r", encoding="utf-8") as f:
        bids = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "추천공고"

    # 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E4057")
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 헤더
    headers = ["번호", "공고명", "발주기관", "추정가격(원)", "계약방법", "입찰마감", "AI점수", "추천이유", "공고URL"]
    col_widths = [5, 40, 20, 15, 12, 18, 8, 40, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 25

    # 데이터
    for i, bid in enumerate(bids, 1):
        eval_info = bid.get("평가", {})
        score = eval_info.get("점수", 0)
        reason = eval_info.get("이유", "")

        row_data = [
            i,
            bid.get("공고명", ""),
            bid.get("발주기관", ""),
            bid.get("추정가격", 0),
            bid.get("계약방법", ""),
            bid.get("입찰마감일시", ""),
            score,
            reason,
            bid.get("공고URL", ""),
        ]

        row_idx = i + 1
        fill = green_fill if score >= 80 else PatternFill("solid", fgColor="FFFDE7")

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.fill = fill
            cell.alignment = center if col in [1, 5, 6, 7] else left

        ws.row_dimensions[row_idx].height = 40

    # 요약 정보
    ws.cell(row=len(bids) + 3, column=1, value=f"총 {len(bids)}건 | 생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    output_file = json_file.replace(".json", ".xlsx")
    wb.save(output_file)
    print(f"✅ 저장: {output_file}")
    return output_file


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        json_to_excel(sys.argv[1])
    else:
        # 가장 최근 recommended_bids JSON 자동 선택
        files = sorted(glob.glob("recommended_bids_*.json"), reverse=True)
        if files:
            print(f"파일 발견: {files[0]}")
            json_to_excel(files[0])
        else:
            print("recommended_bids_*.json 파일을 찾을 수 없습니다.")
