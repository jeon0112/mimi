"""네이버 쇼핑 API를 활용한 품목 단가 자동 조회 에이전트"""

import os
import time
import requests
import openpyxl
from dotenv import load_dotenv

load_dotenv()

NAVER_CLIENT_ID = os.getenv("NAVER_CLIENT_ID")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET")
NAVER_SEARCH_URL = "https://openapi.naver.com/v1/search/shop.json"

DEFAULT_MARGIN = 0.20
DELIVERY_COST = 3000

# 품목 키워드별 최소 합리적 가격 (원)
CATEGORY_MIN_PRICES = {
    "마우스": 8000, "키보드": 10000, "공유기": 25000, "허브": 20000,
    "스피커": 12000, "UTP": 20000, "케이블": 10000, "커넥터": 8000,
    "모니터": 50000, "토너": 8000, "드럼": 15000, "폐토너": 8000,
    "잉크": 5000, "태블릿": 30000, "펜슬": 10000, "노트북": 300000,
    "컴퓨터": 100000, "USB허브": 8000, "SSD": 30000, "메모리": 10000,
    "헤드셋": 10000, "iptime": 20000,
}


def get_category_min(keyword: str) -> int:
    """키워드에서 카테고리 최소가 추출"""
    for cat, min_price in CATEGORY_MIN_PRICES.items():
        if cat in keyword:
            return min_price
    return 500


def search_price(keyword: str, display: int = 10) -> list:
    """네이버 쇼핑에서 키워드로 검색"""
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        print("오류: NAVER_CLIENT_ID 또는 NAVER_CLIENT_SECRET이 설정되지 않았습니다.")
        return []

    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    params = {"query": keyword, "display": display, "sort": "asc"}

    try:
        response = requests.get(NAVER_SEARCH_URL, headers=headers, params=params, timeout=10)
        response.raise_for_status()
        return response.json().get("items", [])
    except Exception as e:
        print(f"  검색 오류 ({keyword}): {e}")
        return []


def extract_prices(items: list, min_threshold: int) -> tuple[list, list]:
    """아이템 목록에서 임계값 이상 가격/제목 추출"""
    prices, titles = [], []
    for item in items:
        try:
            price = int(item.get("lprice", 0))
            if price >= min_threshold:
                prices.append(price)
                titles.append(item.get("title", "").replace("<b>", "").replace("</b>", ""))
        except Exception:
            continue
    return prices, titles


def pick_representative(prices: list, titles: list) -> tuple[int, str]:
    """이상치 제거 후 대표 가격 선택 (평균의 10% 미만 극단값 제거)"""
    if not prices:
        return 0, ""

    sorted_p = sorted(prices)
    avg = sum(sorted_p) / len(sorted_p)

    # 평균의 10% 미만 극단 저가 제거
    filtered = [(p, t) for p, t in zip(sorted_p, sorted(titles, key=lambda x: prices[titles.index(x)] if x in titles else 0))
                if p >= avg * 0.1]
    if not filtered:
        filtered = list(zip(sorted_p, titles))

    f_prices = [p for p, _ in filtered]
    f_titles = [t for _, t in filtered]

    # 하위 1/3 지점 대표가 (최저가와 평균가 사이)
    idx = len(f_prices) // 3
    return f_prices[idx], f_titles[idx] if idx < len(f_titles) else ""


def get_price(keyword: str, fallback_keyword: str = None) -> dict:
    """가격 조회 - 이상 시 키워드 단순화해서 재시도. 상위 업체 목록도 반환."""
    cat_min = get_category_min(keyword)
    items = search_price(keyword)

    prices, titles = extract_prices(items, cat_min)

    # 카테고리 최소가 기준으로 유효 결과 없으면 재시도
    if not prices and fallback_keyword:
        print(f"    → '{fallback_keyword}'로 재검색...")
        items = search_price(fallback_keyword)
        prices, titles = extract_prices(items, cat_min)

    # 카테고리 최솟값이 기본값(500)인 경우에만 전체 결과로 fallback
    if not prices and items and cat_min <= 500:
        all_prices = [int(i.get("lprice", 0)) for i in items if i.get("lprice", 0)]
        all_titles = [i.get("title", "").replace("<b>", "").replace("</b>", "") for i in items]
        if all_prices:
            prices, titles = all_prices, all_titles

    if not prices:
        return {"min_price": None, "title": "", "suppliers": []}

    rep_price, rep_title = pick_representative(prices, titles)

    # 상위 공급업체 3곳 추출 (가격 낮은 순, cat_min 이상인 것)
    valid_items = []
    for item in items:
        try:
            p = int(item.get("lprice", 0))
            if p >= cat_min:
                valid_items.append({
                    "mallName": item.get("mallName", "").replace("<b>", "").replace("</b>", ""),
                    "price": p,
                    "title": item.get("title", "").replace("<b>", "").replace("</b>", "")[:35],
                    "link": item.get("link", ""),
                })
        except Exception:
            continue
    valid_items.sort(key=lambda x: x["price"])
    top_suppliers = valid_items[:3]

    return {"min_price": rep_price, "title": rep_title[:40], "suppliers": top_suppliers}


def calculate_bid_price(cost_price: int, margin: float = DEFAULT_MARGIN) -> int:
    return int((cost_price + DELIVERY_COST) * (1 + margin))


def process_excel(input_file: str, output_file: str = None, margin: float = DEFAULT_MARGIN) -> str:
    if output_file is None:
        output_file = input_file.replace(".xlsx", "_단가조회결과.xlsx")

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row and "품목" in str(row):
            header_row = i
            break

    if not header_row:
        print("헤더를 찾을 수 없습니다.")
        return ""

    headers = [cell.value for cell in ws[header_row]]
    품목_col = headers.index("품목") + 1 if "품목" in headers else 2
    규격_col = headers.index("규격") + 1 if "규격" in headers else 3
    수량_col = headers.index("예정수량") + 1 if "예정수량" in headers else 5
    단가_col = headers.index("단가") + 1 if "단가" in headers else 6
    금액_col = headers.index("금액") + 1 if "금액" in headers else 7

    ws.cell(row=header_row, column=단가_col).value = "단가(입찰)"
    ws.cell(row=header_row, column=금액_col).value = "금액(입찰)"
    매입가_col = ws.max_column + 1
    ws.cell(row=header_row, column=매입가_col).value = "네이버참고가"
    ws.cell(row=header_row, column=매입가_col + 1).value = "검색상품명"

    total_rows = ws.max_row - header_row
    manual_needed = []
    print(f"\n총 {total_rows}개 품목 단가 조회 시작...\n")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        품목 = ws.cell(row=row_idx, column=품목_col).value
        규격 = ws.cell(row=row_idx, column=규격_col).value
        수량 = ws.cell(row=row_idx, column=수량_col).value

        if not 품목:
            continue

        keyword = f"{품목} {규격}".strip() if 규격 else str(품목).strip()
        fallback = str(품목).strip()  # 규격 없이 품목명만으로 재검색

        print(f"  [{row_idx - header_row}/{total_rows}] {keyword} 검색 중...")

        result = get_price(keyword, fallback_keyword=fallback if 규격 else None)
        min_price = result["min_price"]

        if min_price and min_price > 0:
            bid_price = calculate_bid_price(min_price, margin)
            amount = bid_price * int(수량) if 수량 else 0

            ws.cell(row=row_idx, column=단가_col).value = bid_price
            ws.cell(row=row_idx, column=금액_col).value = amount
            ws.cell(row=row_idx, column=매입가_col).value = min_price
            ws.cell(row=row_idx, column=매입가_col + 1).value = result["title"]

            print(f"    참고가: {min_price:,}원 → 입찰단가: {bid_price:,}원 (수량: {수량})")
        else:
            ws.cell(row=row_idx, column=단가_col).value = "수동입력필요"
            manual_needed.append(f"[{row_idx - header_row}] {keyword}")
            print(f"    검색결과 없음 → 수동 입력 필요")

        time.sleep(0.1)

    total_amount = sum(
        ws.cell(row=r, column=금액_col).value or 0
        for r in range(header_row + 1, ws.max_row + 1)
        if isinstance(ws.cell(row=r, column=금액_col).value, (int, float))
    )

    wb.save(output_file)
    print(f"\n✅ 완료! 저장: {output_file}")
    print(f"   총 입찰 예상금액: {total_amount:,}원")

    if manual_needed:
        print(f"\n⚠️  수동 입력 필요 ({len(manual_needed)}건):")
        for item in manual_needed:
            print(f"   {item}")

    return output_file


if __name__ == "__main__":
    import sys

    print("=" * 55)
    print("  품목 단가 자동 조회 에이전트 (네이버 쇼핑)")
    print("=" * 55)

    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = input("\n엑셀 파일 경로를 입력하세요: ").strip().strip('"')

    if not os.path.exists(excel_file):
        print(f"파일을 찾을 수 없습니다: {excel_file}")
        sys.exit(1)

    margin_input = input("마진율 입력 (기본 20%, 엔터 입력 시 기본값): ").strip()
    margin = float(margin_input) / 100 if margin_input else DEFAULT_MARGIN

    process_excel(excel_file, margin=margin)
